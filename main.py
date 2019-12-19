import os
import cv2
import face_recognition
import pickle
from face_recognition.face_recognition_cli import image_files_in_folder
import time
from tkinter import *
from PIL import ImageTk, Image
from tkinter import ttk
from tkinter import messagebox
import datetime
from twilio.rest import Client
import openpyxl
from ttkthemes import ThemedTk

# ----------------------------------------making pkl file of encodings-------------------------------------------
# encodings = face_recognition.face_recognition_cli.scan_known_people('images')
#
# pickle_out = open('encodings.pkl', 'wb')
# pickle.dump(encodings, pickle_out)
# pickle_out.close()


#======================================Defining Main Window=====================================================

root = ThemedTk(theme='radiance')
root.resizable(0, 0)
root.title('Attendance System')
root.iconbitmap('icon.ico')
root.geometry('1000x600')

Login = False
FrontPage_image = ImageTk.PhotoImage(file='attendence.png')
imagetk = ImageTk.PhotoImage(file='recognition.png')
login_page_image = ImageTk.PhotoImage(file='manager.png')


# defining frames


def mainPage():
    root.geometry('1320x620')
    lmain2.destroy()
    wrong.destroy()
    username.destroy()
    username_display.destroy()
    password.destroy()
    password_display.destroy()
    loginBtn.destroy()
    cancelBtn.destroy()

    titleFrame = Frame(root, width='900', height='100')
    titleFrame.pack(side=TOP)

    middleFrame = Frame(root, width='730', height='200')
    middleFrame.place(x=680, y=110)

    picFrame = Frame(root, width='700', height='500')
    picFrame.place(x=40, y=110)

    global lmain
    # imagetk = ImageTk.PhotoImage(file = 'label image.jpg')
    lmain = ttk.Label(picFrame, image=imagetk)
    lmain.grid()

    global bottomFrame
    bottomFrame = Frame(root, width='730', height='100')
    bottomFrame.place(x=680, y=450)

    statusFrame = Frame(root, width='1280', height='10')
    statusFrame.place(x=0, y=600)

    global timerFrame
    timerFrame = Frame(root, width='730', height='100')
    timerFrame.place(x=680, y=300)
    timerFrame.after(1, Run)

    global titleName
    titleName = ttk.Label(titleFrame, text='ATTENDANCE THROUGH FACIAL RECOGNITION',
                          font=('Tw Cen MT', 20))
    titleName.place(x=180, y=20)

    # ----------------------------------------- in middle frame---------------------------------------------------

    # -----------------------taking the user name---------------------------
    global inputName
    inputName = ttk.Label(middleFrame, text='Enter Name:',
                          font=('Tw Cen MT', 13, 'bold'))
    inputName.place(x=60, y=30)

    global name_display
    name_display = ttk.Entry(middleFrame, width=45, font=('Tw Cen MT', 11))
    name_display.place(x=180, y=30)

    # making a drop down menu of courses
    global courseVar
    courseVar = StringVar()
    # courseVar.set('Choose Course')
    global courseName
    courseName = ttk.Label(middleFrame, text='Course:',
                           font=('Tw Cen MT', 13, 'bold'))
    courseName.place(x=60, y=70)

    course_list = ['Choose Course', 'Calculus', 'Physics', 'Fundamentals of Programming', 'English', 'Discrete Maths',
                   'Islamiat']
    global course_menu
    course_menu = ttk.OptionMenu(middleFrame, courseVar, *course_list)
    course_menu.place(x=180, y=70)

    # ========================================Defining status bar=========================================
    # in bottomFrame
    # making a status bar
    global statusbar
    statusbar = ttk.Label(statusFrame, text='Enter Credentials...', font=('calibri(body)', 10, 'italic')
                          , relief=SUNKEN, anchor=W)
    statusbar.grid(row=0, column=0, sticky=W + E)
    statusbar.config(width='1280')

    # -------------------------------------------in timer frame----------------------------------------------------
    global w
    global time_lb
    time_lb = ttk.Label(timerFrame, text='Time Remaining: ',
                        font=('Tw Cen MT', 15, 'bold'))
    time_lb.place(x=60, y=30)
    w = Canvas(timerFrame, width=450, height=450)
    w.place(x=220, y=20)

    global markBtn
    markBtn = ttk.Button(bottomFrame, text='Mark Attendance', command=lambda: main())
    markBtn.grid(row=0, column=1, padx=40, pady=20)

    global resetBtn
    resetBtn = ttk.Button(bottomFrame, text='Reset', command=lambda: reset())
    resetBtn.grid(row=0, column=3, padx=30, pady=20)

    global logoutBtn
    logoutBtn = ttk.Button(bottomFrame, text='Log Out', command=exitit)
    logoutBtn.grid(row=1, column=5, padx=40, pady=20)

    global openBtn
    openBtn = ttk.Button(bottomFrame, text='Check Attendance', command=openAttendence)
    openBtn.grid(row=1, column=1, padx=40, pady=20)

    global sumBtn
    sumBtn = ttk.Button(bottomFrame, text='Summary', command=summary)
    sumBtn.grid(row=1, column=3, padx=40, pady=20)

    global msgBtn
    msgBtn = ttk.Button(bottomFrame, text='Message', command=message)
    msgBtn.grid(row=0, column=5, padx=40, pady=20)


def login():
    root.geometry('1000x600')
    enterBtn.destroy()
    exitBtn.destroy()
    titleName1.destroy()
    titleName2.destroy()
    exitBtn.destroy()
    lmain3.destroy()
    global username
    username = ttk.Label(root, text='Username',
                         font=('Tw Cen MT', 18, 'bold'))
    username.place(x=430, y=210)
    global username_display
    username_display = ttk.Entry(root, width=45, font=('Tw Cen MT', 11))
    username_display.place(x=310, y=250)
    global password
    password = ttk.Label(root, text='Password',
                         font=('Tw Cen MT', 20, 'bold'))
    password.place(x=430, y=290)
    global password_display
    password_display = ttk.Entry(root, show='*', width=45, font=('Tw Cen MT', 11))
    password_display.place(x=310, y=340)

    global lmain2
    lmain2 = Label(root, image=login_page_image)
    lmain2.place(x=430, y=50)

    global wrong
    wrong = Label(root, text='', fg='red', font=('Tw Cen MT', 11))
    wrong.place(x=330, y=440)

    def loginCheck():
        user = username_display.get()
        login_password = password_display.get()
        if user.lower() == 'admin' and login_password.lower() == 'admin':
            global Login
            Login = True
            mainPage()
        else:
            wrong['text'] = 'Credentials entered are not registered. Please try again!'

    global loginBtn
    loginBtn = ttk.Button(root, text='Login', command=loginCheck)
    loginBtn.place(x=420, y=480)

    global cancelBtn
    cancelBtn = ttk.Button(root, text='Cancel', command=entryPage)
    cancelBtn.place(x=420, y=530)


def FrontPage():
    global titleName1
    titleName1 = ttk.Label(root, text='FACIAL RECOGNITION',
                           font=('Tw Cen MT', 40))
    titleName1.place(x=280, y=60)

    global titleName2
    titleName2 = ttk.Label(root, text='ATTENDANCE SYSTEM',
                           font=('Tw Cen MT', 25))
    titleName2.place(x=350, y=120)

    global lmain3
    lmain3 = Label(root, image=FrontPage_image)
    lmain3.place(x=300, y=170)

    global enterBtn
    enterBtn = ttk.Button(root, text='Login', command=login)
    enterBtn.place(x=420, y=480)

    global exitBtn
    exitBtn = ttk.Button(root, text='EXIT', command=root.destroy)
    exitBtn.place(x=420, y=530)


def entryPage():
    root.geometry('1000x600')
    if Login == True or run == True:
        titleName.destroy()
        w.destroy()
        lmain.destroy()
        inputName.destroy()
        courseName.destroy()
        name_display.destroy()
        course_menu.destroy()
        time_lb.destroy()
        statusbar.destroy()
        markBtn.destroy()
        resetBtn.destroy()
        openBtn.destroy()
        msgBtn.destroy()
        sumBtn.destroy()
        logoutBtn.destroy()
    else:
        lmain2.destroy()
        username.destroy()
        username_display.destroy()
        password.destroy()
        password_display.destroy()
        loginBtn.destroy()
        cancelBtn.destroy()

    global titleName1
    titleName1 = ttk.Label(root, text='FACIAL RECOGNITION',
                           font=('Tw Cen MT', 40))
    titleName1.place(x=280, y=40)

    global titleName2
    titleName2 = ttk.Label(root, text='ATTENDANCE SYSTEM',
                           font=('Tw Cen MT', 25))
    titleName2.place(x=350, y=100)

    global enterBtn
    enterBtn = ttk.Button(root, text='Login', command=login)
    enterBtn.place(x=420, y=480)

    global lmain3
    lmain3 = Label(root, image=FrontPage_image)
    lmain3.place(x=300, y=170)

    global exitBtn
    exitBtn = ttk.Button(root, text='EXIT', command=root.destroy)
    exitBtn.place(x=420, y=530)


s = 59
m = 9
h = 0


def Run():
    global s, m, h, w

    # Delete old text
    w.delete('all')
    # Add new text
    if s<10:
        w.create_text(
            [100, 25], anchor=CENTER, text="0%s:0%s:0%ss" % (h, m, s), font=("Tw Cen MT", 20)
        )
    else:
        w.create_text(
            [100, 25], anchor=CENTER, text="0%s:0%s:%ss" % (h, m, s), font=("Tw Cen MT", 20)
        )

    # s+=1

    if m == 0 and s == 0:
        markBtn = ttk.Button(bottomFrame, text='Mark Attendance', state=DISABLED)
        markBtn.grid(row=0, column=1, padx=40, pady=20)
        return
    elif s == 0:
        m -= 1;
        s = 59
    s -= 1
    # After 1 second, call Run again (start an infinite recursive loop)
    timerFrame.after(1000, Run)


# ---------------------------------------------- in top frame-------------------------------------------------


# ----------------------------------------defining the main face recognition function-------------------------------
run = False


def mark():
    global run
    statusbar['text'] = 'Marking Attendance....'
    course = courseVar.get()
    name = student_name

    now = datetime.datetime.now().strftime("%d/%b")
    global wb
    global sheet
    wb = openpyxl.load_workbook('Attendence.xlsx')
    sheet = wb.get_sheet_by_name(course)  # in original program sheet1 will be replaced by
    # varible course already in gui framework.py

    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == name:
            req_row = i
            break

    global req_column

    for i in range(2, sheet.max_column + 1):
        column_value = sheet.cell(row=1, column=i).value
        column_value = column_value.strftime('%d/%b')
        if column_value == now:
            req_column = i
            break

    sheet.cell(row=req_row, column=req_column).value = 'P'
    wb.save('Attendence.xlsx')
    run = True


def main():
    pickle_in = open('encodings.pkl', 'rb')
    pickled_encodings = list(pickle.load(pickle_in))
    pickle_in.close()
    # print(pickled_encodings)

    # taking the user entered name from above gui
    global student_name
    # global name_for_encoding
    course = courseVar.get()
    if course == 'Choose Course':
        messagebox.showinfo('ERROR', 'Course to choose karo yaar')
        return
    student_name = name_display.get()
    statusbar['text'] = 'Checking Name...'

    # taking the index of the particular person entering name

    if student_name in pickled_encodings[0]:
        index_of_name_encodings = pickled_encodings[0].index(student_name)
    else:
        messagebox.showinfo('Error', 'I don\'nt even know who you are!')
        return

    statusbar['text'] = 'Taking Live Image...'
    cam = cv2.VideoCapture(0)
    # cv2.namedWindow("test")
    start = time.time()

    while int(time.time() - start) != 5:
        global frame
        ret, frame = cam.read()
        cv2.imshow("Face Recognizer", frame)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        if not ret:
            break
        k = cv2.waitKey(1)
    encoding_unknown_image = face_recognition.face_encodings(frame)[0]
    # print('the cam:' ,encoding_unknown_image)

    cam.release()
    cv2.destroyAllWindows()

    imgtk = ImageTk.PhotoImage(image=img)
    lmain.imgtk = imgtk
    lmain.configure(image=imgtk)

    # -------------------------if more than one faces shows the number of faces-----------------------------------------#

    # if len(encoding_unknown_image) > 0:
    #     def strength(encoding_unknown_image):
    #         print(f'{len(encoding_unknown_image)} students present in class.')

    # -----------------------------------------------------------------------------------#

    # taking out the particular image and find its encoding and comparing

    image_encoding = pickled_encodings[1][index_of_name_encodings]

    # comparing the results
    statusbar['text'] = 'Validating Data...'
    results = face_recognition.compare_faces([image_encoding], encoding_unknown_image)
    # print(results)

    # now checking both and showing results

    if True in results and student_name in pickled_encodings[0]:
        mark()
        messagebox.showinfo('Face Recognizer', 'Attendance lag gayi hai aapki {} chill karo!'.format(student_name.split(' ')[0]))
        statusbar['text'] = 'showing results'
        reset()
    else:
        # for encoding in pickled_encodings[1]:
        #     result = face_recognition.compare_faces([encoding], encoding_unknown_image)
        #     if result:
        #         index = pickled_encodings[1].index(encoding)
        #         name = pickled_encodings[0][index]
        #         break
        messagebox.showinfo('Face Recognizer', 'You\'re not {}, are you? Proxy lagane nahi doonga!'.format(student_name))
        reset()


# -------------------------Twilio mobile messages-------------------------------------------------------------------#

acc_sid = os.environ.get('ACC_SID')
auth_token = os.environ.get('AUTH_TOKEN')
client = Client(acc_sid, auth_token)


def message():
    client.messages.create(
        to=os.environ.get('MY_PHONE_NUMBER'),
        from_=+18597554541,
        body='\nPlease report to the class or you will not be marked present.'
    )


# ---------------------------------------reset function--------------------------------
def reset():
    name_display.delete(0, END)
    courseVar.set('Choose Course')
    student_name = ''
    statusbar['text'] = 'Enter Credentials'

    imagetk = ImageTk.PhotoImage(file='recognition.png')
    lmain.imagetk = imagetk
    lmain.configure(image=imagetk)


# ---------------------------------------Exit function--------------------------------

def exitit():
    if run:
        for i in range(2, sheet.max_row):
            if sheet.cell(row=i, column=req_column).value == 'P':
                continue
            else:
                sheet.cell(row=i, column=req_column).value = 'A'
        wb.save('Attendence.xlsx')
    entryPage()


def openAttendence():
    os.startfile('Attendence.xlsx')


def summary():
    if run:
        window = ThemedTk(theme='arc')
        window.resizable(0, 0)
        window.geometry('400x400')
        present_list = []

        title_name2 = ttk.Label(window, text='Attendance Summary', font=('Tw Cen MT', 20), style="BW.TLabel").place(
            x=75, y=0)

        for i in range(2, sheet.max_row):
            if sheet.cell(row=i, column=req_column).value == 'P':
                name = sheet.cell(row=i, column=1).value
                present_list.append(name)

        info = ttk.Label(window, text=str(len(present_list)) + '/' + str(sheet.max_row - 1) + ' students are present',
                         font=('Tw Cen MT', 12)).place(x=0, y=50)
        progressbar = ttk.Progressbar(window, orient=VERTICAL, length=220,
                                      value=(len(present_list) / (sheet.max_row - 1)) * 100).place(x=340, y=50)
        percent = round(len(present_list) / (sheet.max_row - 1) * 100, 2)
        percent_label = ttk.Label(window, text='{}%\npresence'.format(percent)).place(x=340, y=290)

        for x in range(0, len(present_list)):
            lmain_x = ttk.Label(window, text=str(x + 1) + '. ' + present_list[x], font=('Tw Cen MT', 12)).place(x=0, y=(
                        80 + x * 25))

        backBtn = ttk.Button(window, text='Return', command=window.destroy).place(x=280, y=350)

        window.mainloop()
    else:
        messagebox.showinfo('Error', 'There\'s nothing to show here yet!')


# ----------------------------------DEFINING BUTTONS AFTER THEIR FUNCTIONS---------------------------------------
FrontPage()

root.mainloop()
