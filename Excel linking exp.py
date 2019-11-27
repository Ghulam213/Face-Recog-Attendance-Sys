import datetime
import openpyxl

def mark():
    name = 'Ghulam'

    now = datetime.datetime.now().strftime("%d/%b")
    global wb
    global sheet
    wb = openpyxl.load_workbook('C:\\Users\\GM\\Desktop\\Book1.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1') #in original program sheet1 will be replaced by
    #varible course already in gui framework.py

    for i in range(1, sheet.max_row+1):
        if sheet.cell(row = i , column = 1).value == name:
            req_row = i
            break

    global req_column
    
    for i in range(2, sheet.max_column+1):
        column_value = sheet.cell(row = 1 , column = i).value
        column_value = column_value.strftime('%d/%b')
        if column_value == now:
            req_column = i
            break
    print(req_row)
    print(req_column)
    sheet.cell(row = req_row , column = req_column).value = 'P'
    wb.save('Book1.xlsx')

def exitit():
    for i in range(2,sheet.max_row):
        if sheet.cell(row = i , column = req_column).value == 'P':
            continue
        else:
            sheet.cell(row = i , column = req_column).value = 'A'
    wb.save('Book1.xlsx')

def main():
    mark()
    exitit()



main()



