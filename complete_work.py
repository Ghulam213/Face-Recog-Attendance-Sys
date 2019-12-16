
import os
import cv2
import face_recognition
import pickle
from face_recognition.face_recognition_cli import image_files_in_folder
import time
from tkinter import *
from PIL import ImageTk, Image
from tkinter import ttk
from tkinter import messagebox
import datetime
from twilio.rest import Client
import openpyxl
from ttkthemes import ThemedTk



#----------------------------------------making pkl file of encodings-------------------------------------------
# encodings = face_recognition.face_recognition_cli.scan_known_people('images')
#
# pickle_out = open('encodings.pkl', 'wb')
# pickle.dump(encodings, pickle_out)
# pickle_out.close()


#======================================Defining Main Window=====================================================

root = ThemedTk(theme='radiance')
root.title('Attendance System')
#root.iconbitmap('icon.ico')
root.geometry('1320x620')

# defining frames

titleFrame = Frame(root, width='900', height='100')
titleFrame.pack(side=TOP)

middleFrame = Frame(root, width='730', height='200')
middleFrame.place(x=680, y=110)

picFrame = Frame(root, width='700', height='500')
picFrame.place(x=40, y=110)
imgtk = ImageTk.PhotoImage(file = 'label image.jpg')
lmain = ttk.Label(picFrame, image = imgtk)
lmain.grid()

bottomFrame = Frame(root, width='730', height='100')
bottomFrame.place(x=680, y=450)

statusFrame = Frame(root, width='1280', height='10')
statusFrame.place(x=0, y=600)

timerFrame = Frame(root, width='730', height='100')
timerFrame.place(x=680, y=300)

# -------------------------------------------in timer frame----------------------------------------------------

time_lb = ttk.Label(timerFrame, text='Time Remaining: ',
                    font=('Tw Cen MT', 15, 'bold'))
time_lb.place(x=60, y=30)
w = Canvas(timerFrame, width=450, height=450)
w.place(x=220, y=20)

s = 59
m = 14
h = 0


def Run():
    global s, m, h

    # Delete old text
    w.delete('all')
    # Add new text
    w.create_text(
        [100, 25], anchor=CENTER, text="%s:%s:%s" % (h, m, s), font=("Tw Cen MT", 25)
    )

    # s+=1

    if m == 0 and s == 0:
        markBtn = Button(bottomFrame, text='Mark Attendance', state=DISABLED)
        markBtn.grid(row=0, column=1, padx=40, pady=20)
        return
    elif s == 0:
        m -= 1;
        s = 59
    s -= 1
    # After 1 second, call Run again (start an infinite recursive loop)
    timerFrame.after(1000, Run)


timerFrame.after(1, Run)

# ---------------------------------------------- in top frame-------------------------------------------------

titleName = ttk.Label(titleFrame, text='ATTENDANCE THROUGH FACIAL RECOGNITION',
                      font=('Tw Cen MT', 30))
titleName.place(x=110, y=20)

# ----------------------------------------- in middle frame---------------------------------------------------

# -----------------------taking the user name---------------------------
inputName = ttk.Label(middleFrame, text='Enter Name:',
                      font=('Tw Cen MT', 13, 'bold'))
inputName.place(x=60, y=30)

name_display = ttk.Entry(middleFrame, width=45, font=('Tw Cen MT', 11))
name_display.place(x=180, y=30)

# making a drop down menu of courses
courseVar = StringVar()
courseVar.set('Choose Course')

courseName = ttk.Label(middleFrame, text='Course:',
                       font=('Tw Cen MT', 13, 'bold'))
courseName.place(x=60, y=70)

course_list = ['Choose Course', 'Calculus', 'Physics', 'Fundamentals of Programming', 'English', 'Discrete Maths',
               'Islamiat']
course_menu = ttk.OptionMenu(middleFrame, courseVar, *course_list)
course_menu.place(x=180, y=70)

# ========================================Defining status bar=========================================
# in bottomFrame
# making a status bar

statusbar = ttk.Label(statusFrame, text='Enter Credentials...', font=('calibri(body)', 10, 'italic')
                      , relief=SUNKEN, anchor=W)
statusbar.grid(row=0, column=0, sticky=W + E)
statusbar.config(width='1280')


# ----------------------------------------defining the main face recognition function-------------------------------
run = False
def mark():
    global run
    statusbar['text'] = 'Marking Attendance....'
    course = courseVar.get()
    name = student_name

    now = datetime.datetime.now().strftime("%d/%b")
    global wb
    global sheet
    wb = openpyxl.load_workbook('Attendence.xlsx')
    sheet = wb.get_sheet_by_name(course)  # in original program sheet1 will be replaced by
    # varible course already in gui framework.py
    print(sheet.max_row + 1)

    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == name:
            req_row = i
            break

    global req_column

    for i in range(2, sheet.max_column + 1):
        column_value = sheet.cell(row=1, column=i).value
        column_value = column_value.strftime('%d/%b')
        if column_value == now:
            req_column = i
            break

    sheet.cell(row=req_row, column=req_column).value = 'P'
    wb.save('Attendence.xlsx')
    run = True


def main():
    pickle_in = open('encodings.pkl', 'rb')
    pickled_encodings = list(pickle.load(pickle_in))
    pickle_in.close()
    # print(pickled_encodings)

    # taking the user entered name from above gui
    global student_name
    # global name_for_encoding
    student_name = name_display.get()
    statusbar['text'] = 'Checking Name...'

    # taking the index of the particular person entering name

    if student_name in pickled_encodings[0]:
        index_of_name_encodings = pickled_encodings[0].index(student_name)
    else:
        messagebox.showinfo('Name Error:', 'Name not found. Please check your name.')

    statusbar['text'] = 'Taking Live Image...'
    cam = cv2.VideoCapture(0)
    cv2.namedWindow("Face Recognizer")
    start = time.time()

    img_counter = 0

    while int(time.time() - start) != 5:
        global frame
        ret, frame = cam.read()
        cv2.imshow("Face Recognizer", frame)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        if not ret:
            break
        k = cv2.waitKey(1)
    global img_name
    # img_name = "{}.jpg".format('live')
    # cv2.imwrite(img_name, frame)
    # print("{} written!".format(img_name))
    encoding_unknown_image = face_recognition.face_encodings(frame)[0]
    #print('the cam:' ,encoding_unknown_image)

    cam.release()
    cv2.destroyAllWindows()

    imgtk = ImageTk.PhotoImage(image=img)
    lmain.imgtk = imgtk
    lmain.configure(image=imgtk)

    #-------------------------if more than one faces shows the number of faces-----------------------------------------#

    # if len(encoding_unknown_image) > 0:
    #     def strength(encoding_unknown_image):
    #         print(f'{len(encoding_unknown_image)} students present in class.')


#-----------------------------------------------------------------------------------#


    # taking out the particular image and find its encoding and comaring

    image_encoding = pickled_encodings[1][index_of_name_encodings]

    # comparing the results
    statusbar['text'] = 'Validating Data...'
    results = face_recognition.compare_faces([image_encoding], encoding_unknown_image)
    # print(results)

    # now checking both and showing results

    if True in results and student_name in pickled_encodings[0]:
        messagebox.showinfo('Face recognitions result', 'Your attendance have been marked successfully!')
        statusbar['text'] = 'showing results'
        mark()
        reset()
    else:
        messagebox.showinfo('Face recognitions result', 'Your Credentials do not match. Try again!')
        reset()


    #-------------------------Twilio mobile messages-------------------------------------------------------------------#

acc_sid = os.environ.get('ACC_SID')
auth_token = os.environ.get('AUTH_TOKEN')

client = Client(acc_sid, auth_token)

def message():
    client.messages.create(
        to=os.environ.get('MY_PHONE_NUMBER'),
        from_=+18597554541,
        body='\nPlease report to the class or you will not be marked present.'
    )


# ---------------------------------------reset function--------------------------------
def reset():
    name_display.delete(0, END)
    courseVar.set('Choose Course')
    student_name = ''
    statusbar['text'] = 'Enter Credentials'
    imgtk = ImageTk.PhotoImage(file='label image.jpg')
    lmain = ttk.Label(picFrame, image=imgtk)
    lmain.grid()

# ---------------------------------------Exit function--------------------------------

def exitit():
    if run:
        for i in range(2,sheet.max_row):
            if sheet.cell(row = i , column = req_column).value == 'P':
                continue
            else:
                sheet.cell(row = i , column = req_column).value = 'A'
        wb.save('Attendence.xlsx')
    root.destroy()

def openAttendence():
    os.startfile('Attendence.xlsx')

def summary():
    if run:
        window = Tk()
        window.title('Summary')
        window.geometry('400x400')
        present_list = []

        title_name2 = Label(window , text = 'Attendance Summary' , font = ('Tw Cen MT',20,'bold')).place(x = 50 , y = 0)
        for i in range(2,sheet.max_row):
            if sheet.cell(row = i , column = req_column).value == 'P':
                name = sheet.cell(row = i , column = 1).value
                present_list.append(name)
        info = Label(window , text = str(len(present_list))+'/'+str(sheet.max_row-2)+' students are present' , font = ('Tw Cen MT',12)).place(x = 0 , y = 30)
        prograssbar = ttk.Progressbar(window , orient = VERTICAL , value = ((len(present_list)/5)*100)).place(x = 350 , y = 50)
        for x in range(0,len(present_list)):
            lmain_x = Label(window , text = str(x+1)+'. '+present_list[x]).place(x = 0 , y = (60 + x*5))
        window.mainloop()
    else:
        messagebox.showinfo('ERROR:','NO INFORMATION TO SHOW!!')


# ----------------------------------DEFINING BUTTONS AFTER THEIR FUNCTIONS---------------------------------------


# making buttons
markBtn = ttk.Button(bottomFrame, text='Mark Attendance', command=lambda: main())
markBtn.grid(row=0, column=1, padx=40, pady=20)

resetBtn = ttk.Button(bottomFrame, text='Reset', command=lambda: reset())
resetBtn.grid(row=0, column=3, padx=30, pady=20)

exitBtn = ttk.Button(bottomFrame, text='EXIT', command=exitit)
exitBtn.grid(row=1, column=5, padx=40, pady=20)

openBtn = ttk.Button(bottomFrame, text='Check Attendance', command=openAttendence)
openBtn.grid(row=1, column=1, padx=40, pady=20)

sumBtn = ttk.Button(bottomFrame, text='Summary', command=summary)
sumBtn.grid(row=1, column=3, padx=40, pady=20)

msgBtn = ttk.Button(bottomFrame, text='Message', command=message)
msgBtn.grid(row=0, column=5, padx=40, pady=20)

root.mainloop()
